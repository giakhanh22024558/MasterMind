# -*- coding: utf-8 -*-
"""cr_to_jira.py — sinh Jira task tree (main + 3 sub-task) từ Gap_Analysis +
Backlog + Acceptance Criteria.

USAGE:
  python cr_to_jira.py <gap.xlsx> <backlog.xlsx> <out_dir>
                       [--no-feature-tag]      # tắt [FEAT-XXX] prefix
                       [--no-cr-tag]           # tắt [CR-XX] prefix
                       [--extra-tag TAG ...]   # thêm custom tag (vd MVP-1, BETA)
  # produces: <out_dir>/cr-XX-task.json + cr-XX-task.md cho mỗi CR đã approve

Tag block trên title:
  [FEAT-XXX] [CR-XX] [<extra...>] <title> — <scope>
  - [FEAT-XXX] mặc định ON; tắt bằng --no-feature-tag
  - [CR-XX] mặc định ON cho CR-derived task; tắt bằng --no-cr-tag
  - Extra tag: --extra-tag MVP-1 --extra-tag BETA → prepend [MVP-1] [BETA]

Requires: pip install openpyxl
"""
import argparse, json, os, re, sys
import openpyxl

PRIORITY_MAP = {"P0": "Highest", "P1": "High", "P2": "Medium"}
SKIP_DECISIONS = {"Invalid / Out-of-scope", "Another Sprint", ""}
ALLOW_DECISIONS = {"This Sprint", "Next Sprint"}
SKIP_LIFECYCLES = {"Done", "Archived"}

def slugify(s):
    s = re.sub(r"[^\w\s-]", "", str(s or "").lower())
    return re.sub(r"[\s_-]+", "-", s).strip("-")

def br2nl(s):
    return str(s or "").replace("<br>", "\n").replace("\r", "")

def split_bullets(s):
    """Tách string thành list bullets (chấp nhận • / - / + / số.)."""
    lines = [ln.strip() for ln in br2nl(s).split("\n") if ln.strip()]
    out = []
    for ln in lines:
        ln = re.sub(r"^[•\-\+\*]\s*", "", ln)
        if ln:
            out.append(ln)
    return out

# --- load Gap_Analysis ---
def load_gap(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Gap Analysis"]
    crs = {}
    # header ở row 1-2 (merged), data từ row 3
    for r in range(3, ws.max_row + 1):
        cid = ws.cell(row=r, column=1).value
        if not cid or not str(cid).startswith("CR-"):
            continue
        crs[cid] = {
            "id":         cid,
            "topic":      ws.cell(row=r, column=2).value or "",
            "criteria":   ws.cell(row=r, column=3).value or "",
            "desc":       ws.cell(row=r, column=4).value or "",
            "as_is":      ws.cell(row=r, column=5).value or "",
            "to_be":      ws.cell(row=r, column=6).value or "",
            "impl_ba":    ws.cell(row=r, column=7).value or "",
            "impl_fe":    ws.cell(row=r, column=8).value or "",
            "impl_be":    ws.cell(row=r, column=9).value or "",
            "est_ba":     ws.cell(row=r, column=10).value or 0,
            "est_fe":     ws.cell(row=r, column=11).value or 0,
            "est_be":     ws.cell(row=r, column=12).value or 0,
            "module":     ws.cell(row=r, column=13).value or "",
            "gap_type":   ws.cell(row=r, column=14).value or "",
            "priority":   ws.cell(row=r, column=15).value or "",
            "decision":   ws.cell(row=r, column=16).value or "",
            "note":       ws.cell(row=r, column=17).value or "",
        }
    return crs

# --- load Backlog -> map CR-XX -> Story ---
def load_backlog(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Backlog"]
    cr_to_story = {}     # CR-XX -> {sid, name, lifecycle, feat_id}
    cur_feat = None
    for r in range(2, ws.max_row + 1):
        fid = ws.cell(row=r, column=3).value
        sid = ws.cell(row=r, column=5).value
        if fid and str(fid).startswith("FEAT-"):
            cur_feat = fid
            continue
        if not sid or not str(sid).startswith("STORY-"):
            continue
        name = ws.cell(row=r, column=6).value or ""
        m = re.match(r"^\[(CR-\d+)\]", name)
        if m:
            cr = m.group(1)
            cr_to_story[cr] = {
                "sid":     sid,
                "name":    name,
                "lifecycle": ws.cell(row=r, column=9).value or "",
                "feat_id": cur_feat,
            }
    return cr_to_story, wb

def load_ac(wb, sid):
    """Return list of (ac_id, text) cho story sid."""
    ac = wb["Acceptance Criteria"]
    result, in_target = [], False
    for r in range(2, ac.max_row + 1):
        s = ac.cell(row=r, column=1).value
        a = ac.cell(row=r, column=2).value
        t = ac.cell(row=r, column=3).value
        if s and not a:
            if in_target: break
            in_target = (s == sid)
        elif in_target and a:
            result.append((a, str(t or "").lstrip("☐ ").strip()))
    return result

# --- build task tree ---
def build_tag_block(cr, story, cfg):
    """Build [TAG1] [TAG2] ... prefix theo config."""
    tags = []
    if cfg["feature_tag"] and story.get("feat_id"):
        tags.append("[%s]" % story["feat_id"])
    if cfg["cr_tag"]:
        tags.append("[%s]" % cr["id"])
    for t in cfg["extra_tags"]:
        tags.append("[%s]" % t)
    return " ".join(tags)

def build_main_title(cr, story, cfg):
    prefix = build_tag_block(cr, story, cfg)
    body = "%s — %s" % (cr["desc"].strip().rstrip("."), cr["criteria"].strip())
    return (prefix + " " + body).strip() if prefix else body

def build_main_desc(cr, ac_list, sid):
    lines = [
        "## Context", "",
        "**As-Is:** " + br2nl(cr["as_is"]),
        "**To-Be:** " + br2nl(cr["to_be"]),
    ]
    if cr["note"]:
        lines.append("**Client Note:** " + br2nl(cr["note"]))
    lines += [
        "**Impacted Module:** " + br2nl(cr["module"]),
        "**Decision:** %s · **Priority:** %s" % (cr["decision"], cr["priority"]),
        "", "## Acceptance Criteria", "",
        "| AC ID | Tiêu chí | Test status |",
        "|---|---|---|",
    ]
    for aid, txt in ac_list:
        lines.append("| %s | ☐ %s | Not tested |" % (aid, txt.replace("|", "/")))
    if not ac_list:
        lines.append("| _(chưa có AC)_ | _Bổ sung sau_ | — |")
    lines += [
        "", "## Implementation breakdown", "",
        "Xem 3 sub-task: `[BA]`, `[FE]`, `[BE]`.",
        "", "## Reference", "",
        "- Backlog: `%s`" % sid,
        "- Gap Analysis: row %s" % cr["id"],
    ]
    return "\n".join(lines)

def build_sub_desc(role, impl_text, est, ac_ids, parent_key):
    bullets = split_bullets(impl_text)
    body = "\n".join("- " + b for b in bullets) if bullets else "Không có công việc cho role này"
    return "\n".join([
        "## Công việc của " + role, "",
        body, "",
        "## Estimation", "",
        "%g man-hours" % float(est or 0), "",
        "## Reference", "",
        "- Main task: " + parent_key,
        "- AC liên quan: " + (", ".join(a for a, _ in ac_ids) if ac_ids else "(chưa có AC)"),
    ])

def build_task_tree(cr, story, ac_list, cfg, project_key="<PROJECT>"):
    main_title = build_main_title(cr, story, cfg)
    ac_ids = [(a, t) for a, t in ac_list]
    labels = ["from-gap-analysis", cr["id"].lower(), slugify(cr["topic"])]
    if cfg["feature_tag"] and story.get("feat_id"):
        labels.append(story["feat_id"].lower())
    for t in cfg["extra_tags"]:
        labels.append(slugify(t))
    main = {
        "fields": {
            "project": {"key": project_key},
            "summary": main_title,
            "issuetype": {"name": "Task"},
            "priority": {"name": PRIORITY_MAP.get(cr["priority"], "Medium")},
            "labels": labels,
            "description": build_main_desc(cr, ac_list, story["sid"]),
            # estimation: tổng — field tuỳ instance, đặt placeholder customfield_10016
            "customfield_10016": float(cr["est_ba"] or 0) + float(cr["est_fe"] or 0) + float(cr["est_be"] or 0),
        }
    }
    subs = []
    for role, impl, est in [
        ("BA", cr["impl_ba"], cr["est_ba"]),
        ("FE", cr["impl_fe"], cr["est_fe"]),
        ("BE", cr["impl_be"], cr["est_be"]),
    ]:
        subs.append({
            "fields": {
                "project": {"key": project_key},
                "summary": "[%s] %s" % (role, main_title),
                "issuetype": {"name": "Sub-task"},
                "parent": {"key": "<MAIN-ISSUE-KEY>"},
                "labels": labels + ["role-" + role.lower()],
                "description": build_sub_desc(role, impl, est, ac_ids, "<MAIN-ISSUE-KEY>"),
                "customfield_10016": float(est or 0),
            }
        })
    return {"main": main, "subs": subs}

def render_md(cr, story, ac_list, tree, cfg):
    out = ["# %s" % build_main_title(cr, story, cfg), "",
           "**Priority:** %s · **Labels:** %s · **Total estimate:** %g man-hours"
           % (PRIORITY_MAP.get(cr["priority"], "Medium"),
              ", ".join(tree["main"]["fields"]["labels"]),
              tree["main"]["fields"]["customfield_10016"]),
           "", tree["main"]["fields"]["description"], "",
           "---", "", "## Sub-tasks", ""]
    for sub in tree["subs"]:
        out += ["### " + sub["fields"]["summary"], "",
                "**Estimate:** %g h" % sub["fields"]["customfield_10016"], "",
                sub["fields"]["description"], "", "---", ""]
    return "\n".join(out)

# --- main ---
def parse_args():
    p = argparse.ArgumentParser(description="Sinh Jira task tree từ Gap + Backlog")
    p.add_argument("gap_xlsx")
    p.add_argument("backlog_xlsx")
    p.add_argument("out_dir")
    p.add_argument("--no-feature-tag", action="store_true", help="Tắt [FEAT-XXX] prefix")
    p.add_argument("--no-cr-tag", action="store_true", help="Tắt [CR-XX] prefix")
    p.add_argument("--extra-tag", action="append", default=[],
                   help="Thêm custom tag (repeat để thêm nhiều, vd: --extra-tag MVP-1 --extra-tag BETA)")
    return p.parse_args()

def main():
    args = parse_args()
    cfg = {
        "feature_tag": not args.no_feature_tag,
        "cr_tag":      not args.no_cr_tag,
        "extra_tags":  list(args.extra_tag),
    }
    os.makedirs(args.out_dir, exist_ok=True)

    crs = load_gap(args.gap_xlsx)
    cr_to_story, backlog_wb = load_backlog(args.backlog_xlsx)

    skipped, generated = [], []
    for cid, cr in sorted(crs.items()):
        if cr["decision"] not in ALLOW_DECISIONS:
            skipped.append((cid, "decision=" + cr["decision"]))
            continue
        story = cr_to_story.get(cid)
        if not story:
            skipped.append((cid, "no backlog story"))
            continue
        if story["lifecycle"] in SKIP_LIFECYCLES:
            skipped.append((cid, "lifecycle=" + story["lifecycle"]))
            continue
        ac_list = load_ac(backlog_wb, story["sid"])
        tree = build_task_tree(cr, story, ac_list, cfg)
        # write
        base = os.path.join(args.out_dir, "%s-task" % cid.lower())
        with open(base + ".json", "w", encoding="utf-8") as f:
            json.dump(tree, f, ensure_ascii=False, indent=2)
        with open(base + ".md", "w", encoding="utf-8") as f:
            f.write(render_md(cr, story, ac_list, tree, cfg))
        generated.append(cid)

    print("Generated: %d CR" % len(generated), generated)
    print("Skipped:   %d CR" % len(skipped))
    for cid, reason in skipped:
        print(" ", cid, "—", reason)

if __name__ == "__main__":
    main()
