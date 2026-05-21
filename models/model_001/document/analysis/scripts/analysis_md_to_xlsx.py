# -*- coding: utf-8 -*-
"""
analysis_md_to_xlsx.py — Sinh sheet .xlsx cho Gap Analysis / Impact Analysis
từ file Markdown nguồn, áp chuẩn style chung (xlsx_style.py).

USAGE:
  python analysis_md_to_xlsx.py gap     <input.md> <output.xlsx>
  python analysis_md_to_xlsx.py impact  <input.md> <output.xlsx>

Quy ước .md nguồn:
  - 1 bảng Markdown duy nhất (bỏ qua heading / text khác).
  - Trong ô: dùng "<br>" để xuống dòng; gạch đầu dòng ghi "• " hoặc "- ".

GAP mode    — bảng phẳng, 1 hàng header, style chuẩn.
IMPACT mode — bảng phẳng 9 cột (CR ID | Impl BA | Impl FE | Impl BE |
              Est BA | Est FE | Est BE | Impacted Module | Decision);
              dựng header 2 tầng có merge (Implementation / Estimation),
              cột Decision có dropdown + tô màu, có dòng TỔNG estimation.

Requires: pip install openpyxl
"""
import sys
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import xlsx_style as X

DECISIONS = ["This Sprint", "Next Sprint", "Another Sprint", "Invalid / Out-of-scope"]
DECISION_FILL = {
    "This Sprint": "C6E0B4", "Next Sprint": "FFE699",
    "Another Sprint": "F8CBAD", "Invalid / Out-of-scope": "D9D9D9",
}


# --- parse markdown table ----------------------------------------------------
def _split_row(line):
    cells = line.strip().split("|")
    if cells and cells[0].strip() == "":
        cells = cells[1:]
    if cells and cells[-1].strip() == "":
        cells = cells[:-1]
    return [c.strip() for c in cells]


def _is_sep(line):
    s = line.strip()
    return s.startswith("|") and set(s) <= set("|-: ")


def parse_first_table(path):
    """Trả về (headers, rows) của bảng Markdown đầu tiên trong file."""
    headers, rows = None, []
    with open(path, encoding="utf-8") as fh:
        for raw in fh:
            line = raw.rstrip("\n")
            if not line.lstrip().startswith("|"):
                if headers is not None:
                    break
                continue
            if _is_sep(line):
                continue
            cells = _split_row(line)
            if headers is None:
                headers = cells
            else:
                rows.append(cells)
    if headers is None:
        raise SystemExit("Không tìm thấy bảng Markdown trong: " + path)
    return headers, rows


def _cell_lines(text):
    """Tách nội dung ô theo <br>; bỏ marker bullet đầu dòng nếu có."""
    parts = [p.strip() for p in str(text).split("<br>")]
    out = []
    for p in parts:
        p = re.sub(r"^[•\-]\s*", "", p)
        out.append(p)
    return out


# --- GAP mode ----------------------------------------------------------------
def build_gap(headers, rows, out):
    wb = Workbook()
    ws = wb.active
    ws.title = "Gap Analysis"
    ncols = len(headers)

    ws.append(headers)
    for row in rows:
        ws.append((list(row) + [""] * ncols)[:ncols])

    X.style_header_row(ws, 1, ncols)
    last = len(rows) + 1
    X.apply_borders(ws, 1, last, ncols)
    for r in range(2, last + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).alignment = X.body_alignment()

    # độ rộng: cột đầu = ID; còn lại chia vừa/rộng
    widths = [X.W_ID] + [X.W_WIDE if i >= 3 else X.W_MED
                         for i in range(2, ncols + 1)]
    X.set_widths(ws, widths)
    X.finish(ws, ncols, last, header_rows=1)
    wb.save(out)
    return len(rows)


# --- IMPACT mode -------------------------------------------------------------
def build_impact(headers, rows, out):
    """
    Bảng phẳng 10 cột:
      CR ID | CR Content (gốc + Note KH) | Impl BA/FE/BE | Est BA/FE/BE
            | Impacted Module | Decision
    -> dựng header 2 tầng (Implementation / Estimation merge).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Impact Analysis"
    NC = 10

    # ----- header 2 tầng -----
    ws.merge_cells("A1:A2"); ws["A1"] = "CR ID"
    ws.merge_cells("B1:B2"); ws["B1"] = "Nội dung CR (gốc + Note KH)"
    ws.merge_cells("C1:E1"); ws["C1"] = "Implementation (nội dung task)"
    ws.merge_cells("F1:H1"); ws["F1"] = "Estimation (man-hours)"
    ws.merge_cells("I1:I2"); ws["I1"] = "Impacted Module"
    ws.merge_cells("J1:J2"); ws["J1"] = "Decision"
    for col, val in [("C2", "BA"), ("D2", "FE"), ("E2", "BE"),
                     ("F2", "BA"), ("G2", "FE"), ("H2", "BE")]:
        ws[col] = val
    for ref in ("A1", "B1", "C1", "F1", "I1", "J1",
                "C2", "D2", "E2", "F2", "G2", "H2"):
        X.style_header_cell(ws[ref])

    def _num(x):
        m = re.search(r"-?\d+(?:\.\d+)?", str(x))
        return float(m.group()) if m else 0

    # ----- data rows -----
    sum_ba = sum_fe = sum_be = 0
    r = 3
    for row in rows:
        row = (list(row) + [""] * NC)[:NC]
        cr, content, ba, fe, be, eba, efe, ebe, module, decision = row
        ws.cell(row=r, column=1, value=cr)
        # CR Content: giữ nguyên cấu trúc, chỉ đổi <br> thành xuống dòng
        ws.cell(row=r, column=2,
                value="\n".join(p.strip() for p in str(content).split("<br>")))
        # implementation: bullet -> nhiều dòng
        for cidx, raw in [(3, ba), (4, fe), (5, be)]:
            lines = _cell_lines(raw)
            ws.cell(row=r, column=cidx,
                    value="\n".join("•  " + ln for ln in lines if ln))
        # estimation
        e_ba, e_fe, e_be = _num(eba), _num(efe), _num(ebe)
        sum_ba += e_ba; sum_fe += e_fe; sum_be += e_be
        ws.cell(row=r, column=6, value=e_ba)
        ws.cell(row=r, column=7, value=e_fe)
        ws.cell(row=r, column=8, value=e_be)
        ws.cell(row=r, column=9, value=module)
        dc = ws.cell(row=r, column=10, value=decision)
        if decision in DECISION_FILL:
            dc.fill = PatternFill("solid", fgColor=DECISION_FILL[decision])
            dc.font = Font(bold=True)
        for c in range(1, NC + 1):
            ws.cell(row=r, column=c).alignment = X.body_alignment(
                center=c in (1, 6, 7, 8, 10))
        ws.cell(row=r, column=1).font = Font(bold=True)
        r += 1

    last_data = r - 1

    # ----- dòng TỔNG -----
    total = r
    ws.merge_cells(start_row=total, start_column=1, end_row=total, end_column=5)
    tc = ws.cell(row=total, column=1, value="TỔNG CỘNG (man-hours)")
    tc.font = Font(bold=True)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    for cidx, val in [(6, sum_ba), (7, sum_fe), (8, sum_be)]:
        cell = ws.cell(row=total, column=cidx, value=val)
        cell.font = Font(bold=True)
        cell.alignment = X.body_alignment(center=True)
        cell.fill = X.total_fill()
    ws.merge_cells(start_row=total, start_column=9, end_row=total, end_column=10)
    gc = ws.cell(row=total, column=9,
                 value=f"Tổng: {sum_ba + sum_fe + sum_be:g} man-hours")
    gc.font = Font(bold=True, color="1F3864")
    gc.alignment = Alignment(horizontal="center", vertical="center")
    gc.fill = X.total_fill()

    # ----- viền + width + freeze + filter -----
    X.apply_borders(ws, 1, total, NC)
    X.set_widths(ws, [X.W_ID, X.W_WIDE, X.W_WIDE, X.W_WIDE, X.W_WIDE,
                      X.W_NARROW, X.W_NARROW, X.W_NARROW, X.W_XWIDE, X.W_SHORT])
    X.finish(ws, NC, last_data, header_rows=2)

    # ----- dropdown Decision qua Combo Box sheet -----
    rng = X.add_combobox_sheet(wb, "Combo Box", "Decision", DECISIONS)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1=rng, allow_blank=True,
                        showErrorMessage=True)
    dv.error = "Chỉ chấp nhận giá trị trong danh sách Decision"
    dv.errorTitle = "Giá trị không hợp lệ"
    dv.add(f"J3:J{last_data}")
    ws.add_data_validation(dv)

    wb.save(out)
    return last_data - 2


# --- main --------------------------------------------------------------------
def main():
    if len(sys.argv) != 4 or sys.argv[1] not in ("gap", "impact"):
        raise SystemExit(__doc__)
    mode, src, dst = sys.argv[1], sys.argv[2], sys.argv[3]
    headers, rows = parse_first_table(src)
    # bỏ dòng tổng nếu có sẵn trong .md (script tự sinh lại)
    rows = [r for r in rows if r and not
            r[0].strip().replace("*", "").upper().startswith(("TỔNG", "TOTAL"))]
    n = build_gap(headers, rows, dst) if mode == "gap" \
        else build_impact(headers, rows, dst)
    print(f"Wrote {dst} — {n} data rows ({mode})")


if __name__ == "__main__":
    main()
