# -*- coding: utf-8 -*-
"""
xlsx_style.py — Chuẩn style .xlsx dùng chung cho mọi sheet sinh ra trong repo.

Đảm bảo MỌI file .xlsx (requirements, features, gap analysis, impact analysis…)
có HÌNH THỨC ĐỒNG NHẤT: cùng màu header, cùng border, cùng cách freeze/filter,
cùng thang độ rộng cột.

Decode từ chuẩn đang dùng (ba_md_to_xlsx.py / Gap Analysis):
  - Header  : nền #1F4E79, chữ trắng in đậm, căn giữa, wrap
  - Body    : căn trên, wrap
  - Border  : viền đơn mảnh #BFBFBF toàn bộ ô
  - freeze hàng header · bật auto-filter

Cài: pip install openpyxl
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================
# HẰNG SỐ CHUẨN
# ============================================================
HEADER_FILL = "1F4E79"      # nền header (xanh đậm)
HEADER_FONT = "FFFFFF"      # chữ header (trắng)
BORDER_COLOR = "BFBFBF"     # màu viền ô
TOTAL_FILL = "BDD7EE"       # nền dòng tổng / nhấn nhẹ

# Thang độ rộng cột chuẩn (đơn vị Excel char-width)
W_ID = 12        # cột mã / ID
W_NARROW = 10    # cột số / ngắn (estimation, count…)
W_SHORT = 16     # cột nhãn ngắn (decision, status…)
W_MED = 24       # cột vừa
W_WIDE = 40      # cột nội dung dài
W_XWIDE = 48     # cột nội dung rất dài (mô tả, module…)


# ============================================================
# STYLE OBJECT
# ============================================================
def header_fill():
    return PatternFill("solid", fgColor=HEADER_FILL)


def header_font():
    return Font(bold=True, color=HEADER_FONT)


def total_fill():
    return PatternFill("solid", fgColor=TOTAL_FILL)


def thin_border():
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)


def header_alignment():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def body_alignment(center=False):
    return Alignment(horizontal="center" if center else "left",
                     vertical="top", wrap_text=True)


# ============================================================
# HÀM TIỆN ÍCH
# ============================================================
def style_header_row(ws, row, ncols):
    """Áp style header cho 1 hàng (1..ncols)."""
    fill, font, align = header_fill(), header_font(), header_alignment()
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def style_header_cell(cell):
    """Áp style header cho 1 ô (dùng cho header merge)."""
    cell.fill = header_fill()
    cell.font = header_font()
    cell.alignment = header_alignment()


def apply_borders(ws, first_row, last_row, ncols):
    """Viền đơn mảnh cho vùng ô."""
    border = thin_border()
    for r in range(first_row, last_row + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = border


def set_widths(ws, widths):
    """Đặt độ rộng cột theo list (dùng thang W_* để nhất quán)."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def finish(ws, ncols, last_data_row, header_rows=1):
    """Freeze hàng header + bật auto-filter."""
    ws.freeze_panes = f"A{header_rows + 1}"
    ws.auto_filter.ref = f"A{header_rows}:{get_column_letter(ncols)}{last_data_row}"


def add_dropdown_inline(ws, col_index, first_row, last_row, values):
    """Dropdown từ danh sách giá trị inline (ít giá trị, ổn định)."""
    col = get_column_letter(col_index)
    dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"',
                        allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)
    dv.add(f"{col}{first_row}:{col}{last_row}")
    return dv


def add_combobox_sheet(wb, title, header, values):
    """
    Tạo sheet 'Combo Box' chứa danh sách giá trị + trả về chuỗi range
    để dùng làm formula1 cho DataValidation (dropdown bền vững, sửa được).
    """
    ws = wb.create_sheet(title)
    h = ws.cell(row=1, column=1, value=header)
    style_header_cell(h)
    border = thin_border()
    for i, v in enumerate(values, start=2):
        cell = ws.cell(row=i, column=1, value=v)
        cell.border = border
    ws.column_dimensions["A"].width = W_MED + 2
    ws.freeze_panes = "A2"
    return f"'{title}'!$A$2:$A${1 + len(values)}"
