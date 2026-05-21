"""
ba_md_to_xlsx.py - generate a pilot .xlsx from a business-analysis table .md.

USAGE:
  python ba_md_to_xlsx.py requirements <input.md> <output.xlsx>
  python ba_md_to_xlsx.py features     <input.md> <output.xlsx>

The .md is the source of truth (context); the .xlsx is the user's manual
pilot copy. Per the Core Rule, write the .xlsx into the working folder's
output/ directory.

  - requirements: the .md has one table per "## <timestamp>" section; the
    .xlsx flattens them into one sheet and prepends a "Run" column.
  - features: the .md has one table; the .xlsx adds dropdowns (Priority,
    In Scope), checkbox-style cells (Ready?, Done?) and merges the
    feature-level cells across each feature's story rows.

Requires: pip install openpyxl
"""
import sys
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("openpyxl is required - install with: pip install openpyxl")

# --- CONFIG -------------------------------------------------------------------
PRIORITY_VALUES = ["Low", "Medium", "High", "Very high"]
IN_SCOPE_VALUES = ["In scope", "Out of scope", "Next phase", "Undecided"]
CHECK_VALUES = ["☐", "☑"]          # unchecked / checked box

HEADER_FILL = "1F4E79"
HEADER_FONT = "FFFFFF"

REQ_HEADERS = ["Req code", "Topic", "Criteria", "Description",
               "Ref. Docs", "Q&A", "Remarks"]
FEATURE_COLS = ["Feature ID", "Feature Name", "Ref. Req (Feature)",
                "Description (Feature)", "Ready?", "Done?", "In Scope"]


# --- MARKDOWN TABLE PARSING ---------------------------------------------------
def _split_row(line):
    cells = line.strip().split("|")
    if cells and cells[0].strip() == "":
        cells = cells[1:]
    if cells and cells[-1].strip() == "":
        cells = cells[:-1]
    return [c.strip() for c in cells]


def _is_separator(line):
    s = line.strip()
    return s.startswith("|") and set(s) <= set("|-: ")


def parse_tables(path):
    """Return a list of (section_title, headers, rows)."""
    tables = []
    state = {"section": "", "headers": None, "rows": []}

    def flush():
        if state["headers"]:
            tables.append((state["section"], state["headers"], state["rows"]))
        state["headers"], state["rows"] = None, []

    with open(path, encoding="utf-8") as fh:
        for raw in fh:
            line = raw.rstrip("\n")
            if line.startswith("## "):
                flush()
                state["section"] = line[3:].strip()
            elif line.lstrip().startswith("|"):
                if _is_separator(line):
                    continue
                cells = _split_row(line)
                if state["headers"] is None:
                    state["headers"] = cells
                else:
                    state["rows"].append(cells)
            elif line.strip() == "":
                continue
            else:
                flush()
    flush()
    return tables


def _pad(row, n):
    return (list(row) + [""] * n)[:n]


# --- XLSX STYLING -------------------------------------------------------------
def _style_header(ws, ncols):
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    font = Font(bold=True, color=HEADER_FONT)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)


def _finish(ws, ncols, nrows, widths):
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, nrows + 2):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if r > 1:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    if nrows:
        ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(ncols), nrows + 1)


def _add_dropdown(ws, col_index, last_row, values):
    col = get_column_letter(col_index)
    dv = DataValidation(type="list",
                        formula1='"' + ",".join(values) + '"',
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("%s2:%s%d" % (col, col, last_row))


# --- BUILDERS -----------------------------------------------------------------
def build_requirements(tables, out):
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"
    body_headers = tables[0][1] if tables else REQ_HEADERS
    headers = ["Run"] + body_headers
    ws.append(headers)
    n = 0
    for section, _hdr, rows in tables:
        for row in rows:
            ws.append([section] + _pad(row, len(body_headers)))
            n += 1
    _style_header(ws, len(headers))
    _finish(ws, len(headers), n, [20, 12, 16, 20, 42, 28, 22, 22])
    wb.save(out)
    return n


def build_features(tables, out):
    if not tables:
        raise SystemExit("No table found in the features .md")
    headers, rows = tables[0][1], tables[0][2]
    ncols = len(headers)
    idx = {h: i + 1 for i, h in enumerate(headers)}

    wb = Workbook()
    ws = wb.active
    ws.title = "Features"
    ws.append(headers)
    for row in rows:
        ws.append(_pad(row, ncols))
    _style_header(ws, ncols)
    _finish(ws, ncols, len(rows),
            [12, 18, 16, 30, 34, 14, 30, 12, 9, 9, 14][:ncols]
            + [16] * max(0, ncols - 11))

    last = len(rows) + 1
    if "Priority" in idx:
        _add_dropdown(ws, idx["Priority"], last, PRIORITY_VALUES)
    if "In Scope" in idx:
        _add_dropdown(ws, idx["In Scope"], last, IN_SCOPE_VALUES)
    if "Ready?" in idx:
        _add_dropdown(ws, idx["Ready?"], last, CHECK_VALUES)
    if "Done?" in idx:
        _add_dropdown(ws, idx["Done?"], last, CHECK_VALUES)

    # merge feature-level cells across each feature's story rows
    fid = idx.get("Feature ID")
    if fid:
        starts = [r for r in range(2, last + 1)
                  if str(ws.cell(row=r, column=fid).value or "").strip()]
        starts.append(last + 1)
        for i in range(len(starts) - 1):
            s, e = starts[i], starts[i + 1] - 1
            if e > s:
                for name in FEATURE_COLS:
                    if name in idx:
                        ci = idx[name]
                        ws.merge_cells(start_row=s, start_column=ci,
                                       end_row=e, end_column=ci)
    wb.save(out)
    return len(rows)


def main():
    if len(sys.argv) != 4 or sys.argv[1] not in ("requirements", "features"):
        raise SystemExit(__doc__)
    mode, src, dst = sys.argv[1], sys.argv[2], sys.argv[3]
    tables = parse_tables(src)
    if mode == "requirements":
        n = build_requirements(tables, dst)
    else:
        n = build_features(tables, dst)
    print("Wrote %s - %d data rows" % (dst, n))


if __name__ == "__main__":
    main()
